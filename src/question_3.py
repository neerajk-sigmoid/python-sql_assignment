import openpyxl
from pathlib import Path
import logging

#setting path in logger
logging.basicConfig(filename='status_update.log',filemode='w',level=logging.DEBUG)
'''class Question_3 is not commented'''
class Question_3:
    flag=True
    #function to import data from created .xlsx file and create table in database using that file
    def solution(self,cursor):
        try:
            self.cursor=cursor
            xlsx_file = Path('Employee compensation.xlsx')
            wb = openpyxl.load_workbook(xlsx_file)
            sh1 = wb['Sheet1']#Sheet1 is the name assigned to the sheet automatically

            #if table created and data is inserted earlier then drop the table
            cursor.execute("Drop table if exists backup_table")
            cursor.execute("""Create table backup_table(Employee_no numeric,
                             Employee_name varchar(250),
                               Department varchar(250),
                               Experience numeric,
                               Total_compensation numeric);""")
            for i in range(2, sh1.max_row + 1):#starting from index 1 as index 0 contains column names
                cursor.execute("Insert into backup_table values('{}','{}','{}','{}','{}')".format(sh1.cell(i, 1).value,
                                                                                                  sh1.cell(i, 2).value,
                                                                                                  sh1.cell(i, 3).value,
                                                                                                  sh1.cell(i, 4).value,
                                                                                                  sh1.cell(i, 5).value))
            logging.info("question 3 executed successfully")

        except Exception as e:
            self.flag=False
            logging.error('error in executing question 3')
            raise Exception(e)
